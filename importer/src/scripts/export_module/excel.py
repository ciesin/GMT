from pathlib import Path

from scripts.export_module.db_constants import *
import scripts.export_module.indicator_constants as ic
import psycopg2.extensions

from scripts.export_module.db_export import db_connect

SHEET_NAME_FIXED_POST = "Fixed Post Health Facilities"
SHEET_NAME_OUTREACHES = "Outreaches"
SHEET_NAME_SETTLEMENTS = "Settlements"

def insert_data_as_table(sheet, data, table_name):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    for row in data:
        sheet.append(row)

    max_row = sheet.max_row
    max_col = sheet.max_column

    # Define the range of data for the table
    table_range = f"A1:{get_column_letter(max_col)}{max_row}"

    # Create a table
    tab = Table(displayName=table_name, ref=table_range)

    # Add a style
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style

    # Add the table to the sheet
    sheet.add_table(tab)

def autowidth_sheet(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width


def append_state_rows(
        conn: psycopg2.extensions.connection,
        data: List[List[Union[str,int]]], 
        type_name: str, ind_name: str, view_name: str):
    with conn.cursor() as cur:
        cur.execute(f"""
                SELECT 
                    '{type_name}', --Type
                    '{ind_name}', --Indicator
                    b1.state_name, --State
                    COUNT(v.global_id) --count non null values
                FROM 
                    {SCHEMA_EXPORT}.{TABLE_STATES} b1
                LEFT JOIN 
                    {SCHEMA_EXPORT}.{view_name} v ON v.b1_guid = b1.state_guid 
                GROUP BY b1.state_name 
                ORDER BY b1.state_name
                """)

        rows = cur.fetchall()

        for r in rows:
            data.append(list(r))

def append_lga_rows(conn: psycopg2.extensions.connection,
                    data: List[List[Union[str,int]]], 
                    type_name: str, ind_name: str, view_name: str):
    with conn.cursor() as cur:
        cur.execute(f"""
                SELECT 
                    '{type_name}', --Type
                    '{ind_name}', --Indicator
                    b2.state_name, --State
                    b2.lga_name, -- LGA 
                    COUNT(v.global_id) --count non null values
                FROM 
                    {SCHEMA_EXPORT}.{TABLE_LGAS} b2                
                LEFT JOIN 
                    {SCHEMA_EXPORT}.{view_name} v ON v.b2_guid = b2.lga_guid 
                GROUP BY b2.state_name, b2.lga_name
                ORDER BY b2.state_name, b2.lga_name
                """)

        rows = cur.fetchall()

        for r in rows:
            data.append(list(r))


def append_ward_rows(
        conn: psycopg2.extensions.connection,
        data: List[List[Union[str,int]]], 
        type_name: str, ind_name: str, view_name: str):
    with conn.cursor() as cur:
        cur.execute(f"""
                SELECT 
                    '{type_name}', --Type
                    '{ind_name}', --Indicator
                    b3.state_name, --State
                    b3.lga_name, -- LGA 
                    b3.ward_name, --Ward
                    COUNT(v.global_id) --count non null values
                FROM 
                    {SCHEMA_EXPORT}.{TABLE_WARDS} b3                
                LEFT JOIN 
                    {SCHEMA_EXPORT}.{view_name} v ON v.b3_guid = b3.ward_guid 
                GROUP BY b3.state_name, b3.lga_name, b3.ward_name
                ORDER BY b3.state_name, b3.lga_name, b3.ward_name
                """)

        rows = cur.fetchall()

        for r in rows:
            data.append(list(r))


def get_state_excel_data(conn: psycopg2.extensions.connection,):
    set_short_name_map = ic.build_set_short_name_map()
    hf_short_name_map = ic.build_hf_short_name_map()
    b_short_name_map = ic.build_b_short_name_map()

    data: List[List[Union[str,int]]] = [
                ['Type', 'Indicator', 'State', 'Count']
            ]

    for ind_name, view_name in set_short_name_map.items():
        append_state_rows(conn, data, 'Settlements', ind_name, view_name)
    for ind_name, view_name in hf_short_name_map.items():
        append_state_rows(conn, data, 'Health Facility', ind_name, view_name)
    for ind_name, view_name in b_short_name_map.items():
        append_state_rows(conn, data, 'Boundary Adjustments', ind_name, view_name)

    return data


def get_lga_excel_data(conn: psycopg2.extensions.connection,):
    set_short_name_map = ic.build_set_short_name_map()
    hf_short_name_map = ic.build_hf_short_name_map()
    b_short_name_map = ic.build_b_short_name_map()

    data: List[List[Union[str,int]]] = [
                ['Type', 'Indicator', 'State', 'LGA', 'Count']
            ]
    for ind_name, view_name in set_short_name_map.items():
        append_lga_rows(conn, data, 'Settlements', ind_name, view_name)
    for ind_name, view_name in hf_short_name_map.items():
        append_lga_rows(conn, data, 'Health Facility', ind_name, view_name)
    for ind_name, view_name in b_short_name_map.items():
        append_lga_rows(conn, data, 'Boundary Adjustments', ind_name, view_name)

    return data


def get_ward_excel_data(conn: psycopg2.extensions.connection,):
    set_short_name_map = ic.build_set_short_name_map()
    hf_short_name_map = ic.build_hf_short_name_map()
    b_short_name_map = ic.build_b_short_name_map()
    data: List[List[Union[str,int]]] = [
                ['Type', 'Indicator', 'State', 'LGA', 'Ward', 'Count']
            ]
    for ind_name, view_name in set_short_name_map.items():
        append_ward_rows(conn, data, 'Settlements', ind_name, view_name)
    for ind_name, view_name in hf_short_name_map.items():
        append_ward_rows(conn, data, 'Health Facility', ind_name, view_name)
    for ind_name, view_name in b_short_name_map.items():
        append_ward_rows(conn, data, 'Boundary Adjustments', ind_name, view_name)

    return data

def export_to_excel(export_dir: Path):
    
    export_indicator_counts(export_dir)
    
    export_after_pilot(export_dir)
    
    
def export_indicator_counts(export_dir: Path):
    from openpyxl import Workbook
    wb = Workbook()

    # grab the active worksheet
    ws_state = wb.active

    assert ws_state is not None

    ws_state.title = "State"
    
    conn, _ = db_connect()

    # Create a new sheet
    ws_lga = wb.create_sheet(title="Lga")

    # Create another new sheet
    ws_ward = wb.create_sheet(title="Ward")

    state_data = get_state_excel_data(conn)

    # print(state_data)

    insert_data_as_table(ws_state, state_data, "State_Table")

    lga_data = get_lga_excel_data(conn)

    insert_data_as_table(ws_lga, lga_data, "LGA_Table")

    ward_data = get_ward_excel_data(conn)

    insert_data_as_table(ws_ward, ward_data, "Ward_Table")

    autowidth_sheet(ws_state)
    autowidth_sheet(ws_lga)
    autowidth_sheet(ws_ward)

    # Save the file
    
    wb.save(str(export_dir / "indicator_counts.xlsx"))    


def get_after_data(
    conn: psycopg2.extensions.connection,
    view_name: str, 
    ) -> List[List]:
    with conn.cursor() as cur:
        cur.execute(f"SELECT * FROM {SCHEMA_EXPORT}.{view_name}")
        assert cur.description is not None
        col_names = [desc[0] for desc in cur.description]
        
        geom_col_idx = col_names.index("geom")
        
        data = cur.fetchall()
        
        del col_names[geom_col_idx]
        
        rows = []
        
        rows.append(col_names)
        
        for d in data:
            row = list(d)
            del row[geom_col_idx]
            rows.append(row)

    return rows


def get_view_comments(
    conn: psycopg2.extensions.connection,
) -> List[List[str]]:
    
    sql = f"""
    SELECT
    c.relname AS "Sheet",
    a.attname AS "Field",
    pgd.description AS "Description"
FROM
    pg_class c
    JOIN pg_attribute a ON a.attrelid = c.oid
    LEFT JOIN pg_description pgd ON pgd.objoid = a.attrelid AND pgd.objsubid = a.attnum
WHERE
    c.relkind = 'v'  -- 'v' stands for views
    AND a.attnum > 0  -- Exclude system columns
    AND c.relname in %s
    AND a.attname != 'geom'
ORDER BY
    c.relname, a.attnum --column order in the view;
    """
    
    with conn.cursor() as cur:
        cur.execute(sql, ((VIEW_HEALTH_FACILITIES_FIXED_POST, VIEW_HEALTH_FACILITIES_OUTREACH, VIEW_SETTLEMENTS),))
        assert cur.description is not None
        col_names = [desc[0] for desc in cur.description]
        
        data = cur.fetchall()
                
        rows = []
        
        rows.append(col_names)
        
        sheet_col_index = 0
        
        for d in data:
            row = list(d)
            if row[sheet_col_index] == VIEW_HEALTH_FACILITIES_FIXED_POST:
                row[sheet_col_index] = SHEET_NAME_FIXED_POST
            elif row[sheet_col_index] == VIEW_HEALTH_FACILITIES_OUTREACH:
                row[sheet_col_index] = SHEET_NAME_OUTREACHES
            elif row[sheet_col_index] == VIEW_SETTLEMENTS:
                row[sheet_col_index] = SHEET_NAME_SETTLEMENTS
                
            rows.append(row)

    return rows
            

def export_after_pilot(export_dir: Path):
    
    conn, _ = db_connect()
    
    from openpyxl import Workbook
    wb = Workbook()
    
    conn, _ = db_connect()
    
    
    # Create a new sheet
    ws_fixed_post = wb.active
    assert ws_fixed_post is not None 
    ws_fixed_post.title = SHEET_NAME_FIXED_POST

    fixed_post_data = get_after_data(conn, VIEW_HEALTH_FACILITIES_FIXED_POST)
        
    insert_data_as_table(ws_fixed_post, fixed_post_data, "FP_HF_Table")

    ws_outreach = wb.create_sheet(title=SHEET_NAME_OUTREACHES)
    outreach_data = get_after_data(conn, VIEW_HEALTH_FACILITIES_OUTREACH)
    insert_data_as_table(ws_outreach, outreach_data, "OUTREACH_Table")
    
    ws_settlement = wb.create_sheet(title="Settlements")
    settlement_data = get_after_data(conn, VIEW_SETTLEMENTS)
    insert_data_as_table(ws_settlement, settlement_data, "SETTLEMENT_Table")
    
    ws_data_dictionary = wb.create_sheet(title="Data Dictionary")
    dict_data = get_view_comments(conn)
    insert_data_as_table(ws_data_dictionary, dict_data, "DATA_DICTIONARY_Table")

    autowidth_sheet(ws_fixed_post)
    autowidth_sheet(ws_outreach)
    autowidth_sheet(ws_settlement)
    autowidth_sheet(ws_data_dictionary)

    # Save the file
    
    wb.save(str(export_dir / "after_pilot.xlsx"))