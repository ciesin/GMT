import openpyxl
import xlrd
from enum import Enum
from pathlib import Path
from typing import List, Optional, Union


class ExcelLibs(Enum):
    OPENPYXL = "openpyxl"
    XLRD = "xlrd"


class CellWrapper:

    def __init__(self, cell, lib: ExcelLibs):
        self.lib = lib
        self.cell = cell

    def get_cell_text(self) -> str:

        if self.lib == ExcelLibs.XLRD:

            if self.cell.ctype == 1:  # text
                return self.cell.value
            elif self.cell.ctype == 0:  # empty
                return ""
            elif self.cell.ctype == 2:  # number
                return str(self.cell.value)
            else:
                raise Exception(f"Unknown cell type: {self.cell.ctype}")

        elif self.lib == ExcelLibs.OPENPYXL:

            v = self.cell.value
            if v is None:
                return ""

            if isinstance(v, str):
                return v
            
            elif isinstance(v, int):
                return str(v)
            
            elif isinstance(v, float):
                return str(v)

            raise Exception(f"Unknown type: {v} {type(v)}")

        raise Exception(f"Unknown lib {self.lib}")


class SheetWrapper:

    def __init__(self, sheet, lib: ExcelLibs):
        self.lib = lib
        self.sheet = sheet

        if self.lib == ExcelLibs.OPENPYXL:
            self.name = sheet.title
            self.nrows = sheet.max_row
            self.ncols = sheet.max_column 
        elif self.lib == ExcelLibs.XLRD:
            self.name = sheet.name
            self.nrows = sheet.nrows
            self.ncols = sheet.ncols
        else:
            raise Exception(f"Unknown lib {self.lib}")

    def row(self, row_index: int) -> List[CellWrapper]:
        """
        :param row_index:  0 based index
        :return: 
        """
        
        if self.lib == ExcelLibs.XLRD:
            return [CellWrapper(c, self.lib) for c in self.sheet.row(row_index)]
        elif self.lib == ExcelLibs.OPENPYXL:
            # 1 based
            row_cells = self.sheet[row_index + 1]
            return [CellWrapper(c, self.lib) for c in row_cells]

        raise Exception(f"Unknown lib {self.lib}")
    
    def cell(self, row_index: int, col_index: int) -> CellWrapper:
        """
        :param row_index: 0 based index
        :param col_index: 0 based index

        :return: 
        """

        if self.lib == ExcelLibs.XLRD:
            #row_cells = self.sheet.row(row_index)
            cell = self.sheet.cell(row_index, col_index)
            return CellWrapper(cell, self.lib) 
        elif self.lib == ExcelLibs.OPENPYXL:
            # 1 based
            cell = self.sheet.cell(row_index + 1, col_index+1)
            return CellWrapper(cell, self.lib)

        raise Exception(f"Unknown lib {self.lib}")


class WorkbookWrapper:
    def __init__(self, file_path: Path):
        self.file_path = file_path

        self._workbook: Optional[Union[xlrd.book.Book, openpyxl.Workbook]] = None

            
    def __enter__(self):
        if self.file_path.suffix.lower() == ".xls":
            # https://stackoverflow.com/questions/33241837/python-xlrd-book-how-to-close-the-files
            self.workbook = xlrd.open_workbook(self.file_path, on_demand=True)
            self.lib = ExcelLibs.XLRD
        elif self.file_path.suffix.lower() == ".xlsx":
            self.lib = ExcelLibs.OPENPYXL
            self.workbook = openpyxl.load_workbook(self.file_path)

        assert self.workbook is not None
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        if not self.workbook:
            return 
        # print(f"CLOSING {self.file_path}")
        if self.lib == ExcelLibs.OPENPYXL:
            self.workbook.close()
        elif self.lib == ExcelLibs.XLRD:
            assert isinstance(self.workbook, xlrd.book.Book)
            # https://stackoverflow.com/questions/33241837/python-xlrd-book-how-to-close-the-files
            self.workbook.release_resources()
            del self.workbook 

    def sheet_by_index(self, sheet_index: int) -> SheetWrapper:
        if self.lib == ExcelLibs.XLRD:
            assert isinstance(self.workbook, xlrd.book.Book)
            return SheetWrapper(                
                self.workbook.sheet_by_index(sheet_index),
                self.lib)

        elif self.lib == ExcelLibs.OPENPYXL:
            return SheetWrapper(
                self.workbook.worksheets[sheet_index],
                self.lib)
        else:
            raise Exception("Unknown lib")
