from copy import copy
import pprint
import openpyxl
import psycopg2
import psycopg2.extensions
import psycopg2.extras
import uuid
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet._write_only import WriteOnlyWorksheet
from pathlib import Path
from typing import List, Optional, Tuple, Dict, cast
from lib import file_utils
import shutil
from openpyxl.styles import PatternFill, Font
from scripts.rew_module.rew_helpers import get_text_value, get_expected_text_value, get_text_value_with_default, get_value, get_value_list
from openpyxl.worksheet.table import Table
from scripts.rew_module.rew_parse import DB_CONSTANTS


NA_VALUE = 'N/A'    

# 1 based row index of header row
HF_CATCHMENTS_SHEET_HEADER_ROW = 6
    


def setup_clean_workbook(lga_name: str) -> Tuple[Workbook, Path]:
    orig_excel_path = Path("/data/rew") / f"{lga_name}_orig.xlsx"
    
    excel_path = Path("/data/rew") / f"{lga_name}.xlsx"
    
    file_utils.remove_file(excel_path)
    
    shutil.copyfile(orig_excel_path, excel_path)
    
    workbook = openpyxl.load_workbook(excel_path)
    
    return workbook, excel_path


def delete_existing_table(hf_catchment_sheet: Worksheet):
    table = None
    for tbl in hf_catchment_sheet.tables.values():
        table = tbl
        print(f"Table name is {table.name}")
        if table.name == "Settlements":
            break

    assert table is not None
    assert table.name == "Settlements"
    
    del hf_catchment_sheet.tables["Settlements"]
    

def setup_header(hf_catchment_sheet: Worksheet) -> Dict[str,int]:
    """
    Adds a new column 'Row Type' and 'Settlement ID (REW)'
    
    :param hf_catchment_sheet: 

    :return: 
    """
    header_row = [ get_text_value_with_default(cast(Cell, c).value) for c in hf_catchment_sheet[HF_CATCHMENTS_SHEET_HEADER_ROW]] 
    
    row_type_col_index = 1
    
    title_row_index = 0
    title_column_index = 1
    header_row.insert(row_type_col_index, 'Row Type')

    # Insert Settlement ID (REW) right before
    # Village/Settlement (REW)
    settlement_id_col_index = header_row.index('Village/Settlement (REW)')
    
    header_row.insert(settlement_id_col_index, 'Settlement ID (REW)')
    
    header_label_to_idx: Dict[str,int] = {value: index for index, value in enumerate(header_row)}
    
    # Preserve the title font
    sheet_title_cell = hf_catchment_sheet.cell(
        # +1 because 1 based
        row=title_row_index+1,             
        column=title_column_index+1)
    sheet_font = copy(sheet_title_cell.font)
    
    # Insert the new columns in the Excel sheet
    hf_catchment_sheet.insert_cols(row_type_col_index+1)
    hf_catchment_sheet.insert_cols(settlement_id_col_index+1)
    
    hf_catchment_sheet.cell(
        HF_CATCHMENTS_SHEET_HEADER_ROW,
        row_type_col_index+1,
        'Row Type'
    )
    hf_catchment_sheet.cell(
        HF_CATCHMENTS_SHEET_HEADER_ROW,
        settlement_id_col_index+1,
        'Settlement ID (REW)'
    )
    
    # put title back
    c = hf_catchment_sheet.cell(
            row=title_row_index+1, 
            column=title_column_index+1, 
            value=sheet_title_cell.value)
    c.font = sheet_font
    
    return header_label_to_idx
 
 

def insert_unmatched_settlements(
    tuple_cur: psycopg2.extensions.cursor, 
    current_fixed_post_row_number: int,
    hf_state_lga_ward_name: Tuple[str,str,str,str], 
    hf_catchment_sheet: Worksheet, 
    row_number: int, 
    header_label_to_idx: Dict[str,int]
    ) -> int:
    """
    
    Note that any matched settlement should be in the catchments tab, see check for
    # Show cases where a settlement is not in the GMT Settlements tab
    """
    
    # return REW data for any settlement that is in not in the catchment
    # meaning we need to add a row
    # note that we have a check
    sql = f"""
    --match the same columns as the inline edit to resuse same function
    SELECT r_set.*,
        r_hf.name as hf_rew_name, 
        r_hf.file_path as hf_rew_file_path,
        r_hf.ri_days,
        True as is_hf_match
        FROM {DB_CONSTANTS.SCHEMA_NAME}.{DB_CONSTANTS.TABLE_HF} r_hf 
        INNER JOIN {DB_CONSTANTS.SCHEMA_NAME}.{DB_CONSTANTS.TABLE_SET} r_set 
            ON r_set.hf_id = r_hf.id
        INNER JOIN {DB_CONSTANTS.SCHEMA_NAME}.{DB_CONSTANTS.TABLE_SET_M2M} r_m2m
            ON r_m2m.set_id = r_set.id
        INNER JOIN partitions.hf_view_latest hf 
            ON hf.global_id = r_hf.global_id 
        inner join boundary.polygon_latest b3 
            on b3.global_id = hf.boundary_polygon
        inner join boundary.polygon_latest b2
            on b2.global_id = b3.boundary_polygon
        inner join boundary.polygon_latest b1
            on b1.global_id = b2.boundary_polygon
    WHERE trim(hf.name) = %(hf_name)s 
        and b3.name = %(hf_ward)s     
        and b2.name = %(hf_lga)s
        and b1.name = %(hf_state)s
        --Non matches are marked with 0 guid   
        AND r_m2m.global_id = '00000000-0000-0000-0000-000000000000'
    """
    
    tuple_cur.execute(sql, {
        # just matching ward is good enough since we should only have r_hf data for the LGA
        # we are processing anyway
        "hf_name": hf_state_lga_ward_name[3],
        "hf_ward": hf_state_lga_ward_name[2],
        "hf_lga": hf_state_lga_ward_name[1],
        "hf_state": hf_state_lga_ward_name[0]
    })
    
    rows = tuple_cur.fetchall()
    
    if len(rows) == 0:
        return row_number
    
    print(f"Inserting {len(rows)} in {hf_state_lga_ward_name[3]}")
    
    fp_line = [ cast(Cell,c).value for c in hf_catchment_sheet[current_fixed_post_row_number]] 
    
    hf_catchment_sheet.insert_rows(row_number, amount=len(rows))
    
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    

    for offset in range(0, len(rows)):
        new_row_number = offset + row_number
        
        rew_set_db_row = rows[offset]
        
        file_name = Path(rew_set_db_row.hf_rew_file_path).name
               
        
        # initialize all values as fp line
        for col_idx, value in enumerate(fp_line, start=1):  # start=1 to match Excel's 1-based indexing
            c = hf_catchment_sheet.cell(row=new_row_number, column=col_idx, value=value)
            c.fill = copy(yellow_fill)
        
        for na_header in [
            'Catchment Population Total (GIS POP)',
            'Catchment Population Total (EST POP)',
            'Catchment Population Total (EST POP + GIS where EST POP is missing)',
            'Catchment Population Inside HF Ward',
            'Catchment Population Outside HF Ward',

        ]:
            hf_catchment_sheet.cell(
                new_row_number, 
                1+header_label_to_idx[na_header],
                NA_VALUE)
       
        fill_in_rew_inline(hf_catchment_sheet, [rew_set_db_row], new_row_number, header_label_to_idx)
        
        c = hf_catchment_sheet.cell(
            new_row_number,
            1+header_label_to_idx['Row Type'],
            "REW Settlement without GMT Match"
        )
        
        c = hf_catchment_sheet.cell(
            new_row_number,
            1+header_label_to_idx['Village/Settlement Name'],
            "Not present in GMT"
        )
    
    new_row_number = row_number + len(rows)
    
    hf_catchment_sheet.cell(
        new_row_number,
        1+header_label_to_idx['Index'],
        new_row_number - HF_CATCHMENTS_SHEET_HEADER_ROW - 1
    )
    
    return new_row_number
    

def handle_all_agg_row(
    cur: psycopg2.extensions.cursor, 
    hf_state_lga_ward_name: Tuple[str,str,str,str], 
    hf_catchment_sheet: Worksheet, 
    row_number: int, 
    header_label_to_idx: Dict[str,int]
    ):
    """
    Handles the GMT row that represents the Health facility (so all outreaches + the fixed post)
    
    :param cur: 
    :param hf_state_lga_ward_name: 
    :param hf_catchment_sheet: 
    :param row_number:  1 based
    :param header_label_to_idx: 
    """    
    
    hf_state, hf_lga, hf_ward, hf_name = hf_state_lga_ward_name
    
    # This HF can be missing so do the N/A values 1st
    
    
    # get the 1st cell to copy the fill color
    ref_cell = hf_catchment_sheet.cell(row_number, 1)
    
                
    for column_name in [        
        # Shouldn't need to do this but doesn't hurt
        'Days of Routine Immunization (fixed post) (REW)',
        # The REW fields are filled in for FP agg row only
        'ANC (REW)', 
        'Family Planning (REW)',
        'Labour & Delivery (REW)']:
        
        c = hf_catchment_sheet.cell(
row_number, 
1+header_label_to_idx[column_name],
NA_VALUE)
    
        c.fill = copy(ref_cell.fill)
    
    # We need the distinct count of REW settlements from other REW Hfs or this one
    # so long as the settlement is in this HFs catchment settlement list
    sql = f"""
    select count(distinct s.id)
    --REW settlements
    from {DB_CONSTANTS.SCHEMA_NAME}.{DB_CONSTANTS.TABLE_SET} s
    inner join {DB_CONSTANTS.SCHEMA_NAME}.{DB_CONSTANTS.TABLE_HF} h
        on s.hf_id = h.id    
    inner join {DB_CONSTANTS.SCHEMA_NAME}.{DB_CONSTANTS.TABLE_SET_M2M} m2m
        on m2m.set_id = s.id
    --that are in this HFs catchment
    inner join partitions.sn_view_latest sn
        on sn.global_id = m2m.global_id
            and sn.is_primary
    inner join partitions.sp_view_latest sp
        on sp.global_id = sn.settlement_part
    inner join partitions.ci_view_latest ci
        on ci.settlement_part = sp.global_id
            and ci.type != 'exclude'
    inner join partitions.hf_view_latest hf_maybe_out
        on hf_maybe_out.global_id = ci.health_facility_point
    inner join partitions.hf_view_latest hf
        on case when hf_maybe_out.type = 'fixed_post' then
            hf_maybe_out.global_id = hf.global_id
        else
            hf_maybe_out.parent = hf.global_id
        end
    inner join boundary.polygon_latest b3
        on b3.global_id = hf.boundary_polygon
    inner join boundary.polygon_latest b2
        on b2.global_id = b3.boundary_polygon
    inner join boundary.polygon_latest b1
        on b1.global_id = b2.boundary_polygon
    WHERE b2.name = %(lga_name)s
        and b1.name = %(state_name)s        
        and trim(hf.name) = %(hf_name)s
"""
    dict_args = {
        'state_name': hf_state,
        'lga_name': hf_lga,
        'hf_name': hf_name,
        'ward_name': hf_ward
    }
    cur.execute(sql, dict_args)
    rows = cur.fetchone()
    rew_count_matched = rows[0]
    
    # These are the inserted REW settlement count
    sql = f"""
select count(distinct s.id)
    --REW settlements
    from {DB_CONSTANTS.SCHEMA_NAME}.{DB_CONSTANTS.TABLE_SET} s
    inner join {DB_CONSTANTS.SCHEMA_NAME}.{DB_CONSTANTS.TABLE_HF} h
        on s.hf_id = h.id    
    inner join {DB_CONSTANTS.SCHEMA_NAME}.{DB_CONSTANTS.TABLE_SET_M2M} m2m
        on m2m.set_id = s.id
    inner join partitions.hf_view_latest hf
        on hf.global_id = h.global_id
    inner join boundary.polygon_latest b3
        on b3.global_id = hf.boundary_polygon
    inner join boundary.polygon_latest b2
        on b2.global_id = b3.boundary_polygon
    inner join boundary.polygon_latest b1
        on b1.global_id = b2.boundary_polygon
    where m2m.global_id = '00000000-0000-0000-0000-000000000000'
        AND b2.name = %(lga_name)s
        and b1.name = %(state_name)s        
        and trim(hf.name) = %(hf_name)s
    """
    
    cur.execute(sql, dict_args)
    rows = cur.fetchone()
    rew_count_unmatched = rows[0]
    
    # Query to find overall total of all REW settlements whether they matched 
    # this GMT HF or not
    sql = f"""
    
select 
    sum(s.total_pop), 
    h.id
from {DB_CONSTANTS.SCHEMA_NAME}.{DB_CONSTANTS.TABLE_HF} h
inner join {DB_CONSTANTS.SCHEMA_NAME}.{DB_CONSTANTS.TABLE_SET} s 
    on s.hf_id = h.id
--Assume we have them all matched
inner join partitions.hf_view_latest hf on hf.global_id = h.global_id 
inner join boundary.polygon_latest b3 on b3.global_id = hf.boundary_polygon
WHERE 
    trim(hf.name) = %(hf_name)s 
    and b3.name = %(ward_name)s
    --GMT hf lga/state should match REW hf gmt lga/state
    --We have a row count == 1 check 
    and h.gmt_lga = %(lga_name)s
GROUP BY h.id
    """
    
    cur.execute(sql, dict_args)
    
    rows = cur.fetchall()
    
    if len(rows) == 0:
        print(f"GMT HF {hf_name} NOT found in REW")
        
        c = hf_catchment_sheet.cell(
            row_number,
            1+header_label_to_idx['Row Type'],
            "Health Facility (FP + Outreach) -- No REW Sheet Found!"
        )
        
        c.fill = copy(ref_cell.fill)
            
        c = hf_catchment_sheet.cell(
            row_number, 
            1+header_label_to_idx['Catchment Population Total (REW)'],
            "No REW Sheet Found!")
        
        c.fill = copy(ref_cell.fill)
        
        c = hf_catchment_sheet.cell(
            row_number, 
            1+header_label_to_idx['Settlement ID (REW)'],
            # distinct count
            rew_count_matched+rew_count_unmatched)
        
        c.fill = copy(ref_cell.fill)
        
        return 
    
    if len(rows) != 1:
        raise Exception(f"Expected 1 match for {hf_name}")
    
    total_pop = rows[0][0]
    
    c = hf_catchment_sheet.cell(
        row_number,
        1+header_label_to_idx['Row Type'],
        "Health Facility (FP + Outreach)"
    )
    
    c.fill = copy(ref_cell.fill)
        
    c = hf_catchment_sheet.cell(
row_number, 
1+header_label_to_idx['Catchment Population Total (REW)'],
get_value(total_pop, "No Sum in Target Population sheet"))
    
    c.fill = copy(ref_cell.fill)
    
    c = hf_catchment_sheet.cell(
            row_number, 
            1+header_label_to_idx['Settlement ID (REW)'],
            rew_count_matched+rew_count_unmatched)
    
    c.fill = copy(ref_cell.fill)
    
    
def handle_mobile_agg_row(
    cur: psycopg2.extensions.cursor, 
    hf_state_lga_ward_name: Tuple[str,str,str,str], 
    hf_catchment_sheet: Worksheet, 
    row_number: int, 
    header_label_to_idx: Dict[str,int]
    ):
    """
    :param cur: 
    :param hf_state_lga_ward_name:  GMT boundary names + TRIMMED GMT hf name 
    :param hf_catchment_sheet: 
    :param row_number: 1 based row number
    :param header_label_to_idx: label to 0 based index
    """    
    
    hf_state, hf_lga, hf_ward, hf_name = hf_state_lga_ward_name
    
    assert hf_name == 'Mobile'
    
    # We need the distinct count of everything in the ward
    # but in any catchment
    sql = f"""    
with gmt_ward_sns AS (
    SELECT
        sn.global_id
    FROM partitions.sn_view_latest sn          
    inner join boundary.polygon_latest b3 
        on b3.global_id = sn.boundary_polygon
    inner join boundary.polygon_latest b2 
        on b2.global_id = b3.boundary_polygon
    inner join boundary.polygon_latest b1 
        on b1.global_id = b2.boundary_polygon
    WHERE b2.name = %(lga_name)s
        and b1.name = %(state_name)s
        AND sn.is_primary
        AND b3.name = %(ward_name)s
),
--Any settlement in a catchment in this Wards HFs
gmt_ward_catch_sns AS (    
    select 
        sn.global_id,
        SUM(sp.computed_pop * ci.population_perc / 100) AS gis_pop,
        SUM(COALESCE(sn.estimated_pop,0) * ci.population_perc / 100) AS est_pop,
        SUM(COALESCE(sn.estimated_pop, sp.computed_pop) * ci.population_perc / 100) AS est_gis_pop 
    from partitions.sn_view_latest sn 
    inner join boundary.polygon_latest b3 
        on b3.global_id = sn.boundary_polygon
    inner join boundary.polygon_latest b2 
        on b2.global_id = b3.boundary_polygon
    inner join boundary.polygon_latest b1 
        on b1.global_id = b2.boundary_polygon
    inner join partitions.sp_view_latest sp 
        on sp.global_id = sn.settlement_part
    inner join partitions.ci_view_latest ci 
        on ci.settlement_part = sp.global_id             
        and ci.type = 'generated'    
    WHERE --for perf also check ward
        b2.name = %(lga_name)s
        and b1.name = %(state_name)s
        and sn.is_primary
        AND b3.name = %(ward_name)s
    GROUP BY sn.global_id
)
select count(distinct s.id)
    --REW settlements
    from {DB_CONSTANTS.SCHEMA_NAME}.{DB_CONSTANTS.TABLE_SET} s
    inner join {DB_CONSTANTS.SCHEMA_NAME}.{DB_CONSTANTS.TABLE_HF} h
        on s.hf_id = h.id    
    inner join {DB_CONSTANTS.SCHEMA_NAME}.{DB_CONSTANTS.TABLE_SET_M2M} m2m
        on m2m.set_id = s.id    
    inner join partitions.sn_view_latest sn
        on sn.global_id = m2m.global_id
            and sn.is_primary
    inner join partitions.sp_view_latest sp 
        on sp.global_id = sn.settlement_part
    inner join gmt_ward_sns on
        gmt_ward_sns.global_id = sn.global_id
    left join gmt_ward_catch_sns c
        on c.global_id = sn.global_id
    WHERE --Uncovered pop >= 0.5, either est. or computed
    (
        0.5 <=
        (
            --uncovered estimated pop
            COALESCE(sn.estimated_pop,0) - COALESCE(c.est_pop, 0)
        )
        OR        
        0.5 <=
        (
            --uncovered computed pop
            COALESCE(sp.computed_pop,0) - COALESCE(c.gis_pop, 0)
        )
    )
"""
    dict_args = {
        'state_name': hf_state,
        'lga_name': hf_lga,
        'hf_name': hf_name,
        'ward_name': hf_ward
    }
    cur.execute(sql, dict_args)
    rows = cur.fetchone()
    rew_count_matched = rows[0]
    
    # get the 1st cell to copy the fill color
    ref_cell = hf_catchment_sheet.cell(row_number, 1)
    
    c = hf_catchment_sheet.cell(
        row_number, 
        1+header_label_to_idx['Settlement ID (REW)'],
        # distinct count, only matched because unmatched are always under a GMT HF
        rew_count_matched)
    
    c.fill = copy(ref_cell.fill)


def handle_fixed_post_agg_row(
    cur: psycopg2.extensions.cursor, 
    hf_state_lga_ward_name: Tuple[str,str,str,str], 
    hf_catchment_sheet: Worksheet, 
    row_number: int, 
    header_label_to_idx: Dict[str,int]
    ):
    """
    :param cur: 
    :param hf_state_lga_ward_name:  GMT boundary names + TRIMMED GMT hf name 
    :param hf_catchment_sheet: 
    :param row_number: 1 based row number
    :param header_label_to_idx: label to 0 based index
    """    
    
    hf_ward = hf_state_lga_ward_name[2]
    hf_name = hf_state_lga_ward_name[3]
    
    sql = f"""    
select h.id, h.ri_days, h.services, h.services_yes_no
from {DB_CONSTANTS.SCHEMA_NAME}.{DB_CONSTANTS.TABLE_HF} h
inner join partitions.hf_view_latest hf on hf.global_id = h.global_id 
inner join boundary.polygon_latest b3 on b3.global_id = hf.boundary_polygon
where trim(hf.name) = %s and b3.name = %s;

    """
    
    cur.execute(sql, (hf_name, hf_ward ))
    
    rows = cur.fetchall()    
    
    # get the 1st cell to copy the fill color
    ref_cell = hf_catchment_sheet.cell(row_number, 1)
    
    # do na values 1st that apply in both the match and not match case
    for column_name in [
        'Settlement ID (REW)',
        'Catchment Population Total (REW)']:
        
        c = hf_catchment_sheet.cell(
row_number, 
1+header_label_to_idx[column_name],
NA_VALUE)
    
        c.fill = copy(ref_cell.fill)
    
    if len(rows) == 0:
        print(f"GMT HF {hf_name} NOT found in REW")
        c = hf_catchment_sheet.cell(
            row_number,
            1+header_label_to_idx['Row Type'],
            "Health Facility (FP Only) -- No REW Sheet Found!"
        )
        
        c.fill = copy(ref_cell.fill)
        
        no_value = "No REW Sheet Found!"
        
        for column_name in [
            'Days of Routine Immunization (fixed post) (REW)',
            'ANC (REW)',
            'Family Planning (REW)',
            'Labour & Delivery (REW)'
        ]:
            c = hf_catchment_sheet.cell(
                row_number,
                1+header_label_to_idx[column_name],
                no_value
            )
            
            c.fill = copy(ref_cell.fill)
        
        return 
    
    if len(rows) != 1:
        raise Exception(f"Expected 1 match for {hf_name}")
    
    ri_days = rows[0][1]
    
    services = rows[0][2]
    services_yes_no = rows[0][3]
    
    anc = 'Unknown, not in background and services tab'
    family = anc 
    labour = anc 
    
    for i, serv in enumerate(services):
        if serv.lower().strip() == "Antenatal Care".lower().strip():
            anc = services_yes_no[i]
        elif serv.lower().strip() == "Delivery".lower().strip():
            labour = services_yes_no[i]
        elif serv.lower().strip() == "Family Planning".lower().strip():
            family = services_yes_no[i]
    
    c = hf_catchment_sheet.cell(
        row_number,
        1+header_label_to_idx['Row Type'],
        "Health Facility (FP Only)"
    )
    
    c.fill = copy(ref_cell.fill)
        
    
    c = hf_catchment_sheet.cell(
row_number, 
1+header_label_to_idx['Days of Routine Immunization (fixed post) (REW)'],
get_value(ri_days, "No RI days info available in Background and Services sheet"))
    
    c.fill = copy(ref_cell.fill)
    
    c = hf_catchment_sheet.cell(
row_number, 
1+header_label_to_idx['ANC (REW)'],
anc)
    
    c.fill = copy(ref_cell.fill)
    
    c = hf_catchment_sheet.cell(
row_number, 
1+header_label_to_idx['Family Planning (REW)'],
family)
    
    c.fill = copy(ref_cell.fill)
    
    c = hf_catchment_sheet.cell(
row_number, 
1+header_label_to_idx['Labour & Delivery (REW)'],
labour)
    
    c.fill = copy(ref_cell.fill)
    
    

def fix_mobile_rows(hf_catchment_sheet: Worksheet, set_name: Optional[str], row_number: int, header_label_to_idx: Dict[str, int]):
    
    # get the 1st cell to copy the fill color
    ref_cell = hf_catchment_sheet.cell(row_number, 1)
    
    for col_name in [
        'Latitude', 'Longitude'
    ]:
        c = hf_catchment_sheet.cell(
    row_number, 
    1+header_label_to_idx[col_name],
    NA_VALUE)
    
        c.fill = copy(ref_cell.fill)
        
    is_agg_row = set_name in [NA_VALUE, ""] or set_name is None
    
    
    
    if not is_agg_row:
        c = hf_catchment_sheet.cell(
            row_number,
            1+header_label_to_idx['Row Type'],
            "Mobile Settlement"
        )
        c.fill = copy(ref_cell.fill)
        return 
    
    c = hf_catchment_sheet.cell(
        row_number,
        1+header_label_to_idx['Row Type'],
        "Mobile Ward Totals"
    )
    c.fill = copy(ref_cell.fill)
    
    for col_name in [
        'ANC (REW)', 'Family Planning (REW)', 'Labour & Delivery (REW)',
        'Catchment Population Total (REW)'
    ]:
        c = hf_catchment_sheet.cell(
    row_number, 
    1+header_label_to_idx[col_name],
    NA_VALUE)
    
        c.fill = copy(ref_cell.fill)


def handle_outreach_agg_row(
    hf_catchment_sheet: Worksheet, 
    row_number: int, 
    header_label_to_idx: Dict[str,int]
    ):
    
    # get the 1st cell to copy the fill color
    ref_cell = hf_catchment_sheet.cell(row_number, 1)
        
    for column_name in [
        'Settlement ID (REW)',
        # Because not clear how to calculate only this GMT outreaches pop
        'Catchment Population Total (REW)',
        # The REW fields are filled in for FP agg row only
        'ANC (REW)', 
        'Family Planning (REW)',
        'Labour & Delivery (REW)'    
    ]:
        
        c = hf_catchment_sheet.cell(
row_number, 
1+header_label_to_idx[column_name],
NA_VALUE)
    
        c.fill = copy(ref_cell.fill)
        
    
    c = hf_catchment_sheet.cell(
        row_number,
        1+header_label_to_idx['Row Type'],
        "Health Facility (Outreach)"
    )
    
    c.fill = copy(ref_cell.fill)
    
    
def get_sn_matches(cur, set_name: str, set_ward: str, hf_line: List, header_label_to_idx: Dict[str, int]):
    
    lat_lon_tolerance = 1e-6
    set_lat = hf_line[header_label_to_idx['Latitude (Settlement)']]
    set_lon = hf_line[header_label_to_idx['Longitude (Settlement)']]
        
    cur_args = {
        'set_name': set_name,
        'ward_name': set_ward,
        'tolerance': lat_lon_tolerance,
        'lat': set_lat,
        'lon': set_lon 
    }
    
    cur.execute(f"""
SELECT sn.global_id 
FROM partitions.sn_view_latest sn 
inner join boundary.polygon_latest b3 on b3.global_id = sn.boundary_polygon
--Note GMT name can have whitespace
WHERE trim(sn.name) = %(set_name)s AND b3.name = %(ward_name)s
and sn.is_primary
--Can have many GMT names in same ward with same name, so we also look at lat/lon
AND ABS(ST_X(sn.geom) - %(lon)s) < %(tolerance)s
AND ABS(ST_Y(sn.geom) - %(lat)s) < %(tolerance)s
""", cur_args)
        
    sn_matches = cur.fetchall()
        
    if len(sn_matches) != 1:
        print(sn_matches)
        pprint.pp(cur_args)
        raise Exception("Expect 1 match")    
    
    return sn_matches


def get_rew_rows(
    tuple_cur: psycopg2.extensions.cursor, 
    sn_guid: str,     
    # Used to identify the GMT HF
    hf_state_lga_ward_name: Tuple[str,str,str,str],) -> List:
    """
    For a given Excel GMT row, we need to fetch the corresponding REW Settlement
    
    Note that the hf of the REW might not match    
    
    """
    
    # Find the GMT Fixed post Health facility for this line
    if hf_state_lga_ward_name[3] == "Mobile":
        gmt_hf_guid = str(uuid.UUID(int=0))
    else:
        tuple_cur.execute(f"""
    SELECT hf.global_id 
        from partitions.hf_view_latest hf
        inner join boundary.polygon_latest b3 on b3.global_id = hf.boundary_polygon
        inner join boundary.polygon_latest b2 on b2.global_id = b3.boundary_polygon
        inner join boundary.polygon_latest b1 on b1.global_id = b2.boundary_polygon
        where b1.name = %s and b2.name = %s and b3.name = %s
        and trim(hf.name) = %s
        and hf.type = 'fixed_post'               
        """, hf_state_lga_ward_name)
        
        
        gmt_hf = tuple_cur.fetchall()
        
        if len(gmt_hf) != 1:
            raise Exception(f"Could not find 1 GMT hf for {hf_state_lga_ward_name}")
        
        gmt_hf_guid = gmt_hf[0].global_id
    
    # get all REW matches for snguid
    tuple_cur.execute(f"""
SELECT 
    s.*, 
    h.name as hf_rew_name, 
    h.file_path as hf_rew_file_path,
    h.ri_days,
    --the REW could be for another HF
    gmt_hf.global_id = %(gmt_hf_guid)s AS is_hf_match
FROM 
{DB_CONSTANTS.SCHEMA_NAME}.{DB_CONSTANTS.TABLE_SET} s
INNER JOIN {DB_CONSTANTS.SCHEMA_NAME}.{DB_CONSTANTS.TABLE_SET_M2M} m2m 
    ON m2m.set_id = s.id
INNER JOIN {DB_CONSTANTS.SCHEMA_NAME}.{DB_CONSTANTS.TABLE_HF} h
    ON h.id = s.hf_id
INNER JOIN partitions.hf_view_latest gmt_hf 
    ON gmt_hf.global_id = h.global_id
WHERE m2m.global_id = %(sn_guid)s 
--put matches 1st for this HF, but we are NOT filtering, we want any HF 
ORDER BY gmt_hf.global_id = %(gmt_hf_guid)s DESC
                    """, {'gmt_hf_guid': gmt_hf_guid, 'sn_guid': sn_guid})
        
    rows = tuple_cur.fetchall()
    
    if len(rows) == 0:
        #print(f"Number of matches is {len(rows)}, skipping")
        return [] 
    else:        
        return rows

def fill_in_rew_inline(
    hf_catchment_sheet: Worksheet, 
    rew_set_db_rows: List, 
    row_number: int, 
    header_label_to_idx: Dict[str, int]):
    
    # for not new row, mismatch color should be yellow
    assert len(rew_set_db_rows) > 0
    rew_set_db_row = rew_set_db_rows[0]
    
    file_name = Path(rew_set_db_row.hf_rew_file_path).name
    no_value_target_pop =  f"Not Present in REW 'Target Population' sheet in {file_name}"
    no_value_catchment_area = f"Not Present in REW 'Catchment Area for Services' sheet in {file_name}"
    no_value_background_services = f"Not Present in REW 'Background and Services' sheet in {file_name}"
    
    assert len(rew_set_db_row.hf_rew_name) > 0
    
    assert not isinstance(hf_catchment_sheet, WriteOnlyWorksheet)
    
    hf_catchment_sheet.cell(
        row_number,
        1+header_label_to_idx['Index'],
        row_number - HF_CATCHMENTS_SHEET_HEADER_ROW - 1
    )
    
    hf_catchment_sheet.cell(
        row_number, 
        1+header_label_to_idx['Health Facility (REW)'],
        ", ".join([r.hf_rew_name for r in rew_set_db_rows]))
    
    assert len(rew_set_db_row.name) > 0
    
    hf_catchment_sheet.cell(
        row_number, 
        1+header_label_to_idx['Village/Settlement (REW)'],
        ", ".join([r.name for r in rew_set_db_rows]))
    
    hf_catchment_sheet.cell(
        row_number, 
        1+header_label_to_idx['Settlement ID (REW)'],
        ", ".join([f"rew_db_id_{r.id}" for r in rew_set_db_rows]))
    
    hf_catchment_sheet.cell(
        row_number, 
        1+header_label_to_idx['Settlement Type (Urban/ Rural) (REW)'],
        ", ".join(get_value_list([r.type for r in rew_set_db_rows], no_value_target_pop))
    )
    
    
    # take first value with len > 0
    
    fixed_outreach_mobile = no_value_catchment_area
    
    fixed_outreach_mobile_list = []
    
    for r in rew_set_db_rows:
        fixed_outreach_mobile = ""
        
        fp_value = cast(str, get_value(r.dist_fp, ""))
        outreach_value = cast(str, get_value(r.dist_out, ""))
        mobile_value = cast(str, get_value(r.dist_mobile, ""))
    
        if len(fp_value) > 0:
            fixed_outreach_mobile = f"Fixed ({fp_value})"
        elif len(outreach_value) > 0:
            fixed_outreach_mobile = f"Outreach ({outreach_value})"
        elif len(mobile_value) > 0:
            fixed_outreach_mobile = f"Mobile ({mobile_value})"
            
        fixed_outreach_mobile_list.append(fixed_outreach_mobile)
    
    hf_catchment_sheet.cell(
        row_number, 
        1+header_label_to_idx['Fixed/Outreach/Mobile (REW)'],
        ", ".join(get_value_list(fixed_outreach_mobile_list, no_value_catchment_area))
    )
                
    hf_catchment_sheet.cell(
        row_number, 
        1+header_label_to_idx['Type of Immunization Sessions (FS, OS1, OS2, etc.) (REW)'],
        ", ".join(get_value_list([r.sessions for r in rew_set_db_rows], no_value_catchment_area))
    )
    
    hf_catchment_sheet.cell(
        row_number, 
        1+header_label_to_idx['Days of Routine Immunization (fixed post) (REW)'],
        ", ".join(get_value_list([r.ri_days for r in rew_set_db_rows], no_value_background_services))
    )
    
    hf_catchment_sheet.cell(
        row_number, 
        1+header_label_to_idx['Total Population (REW)'],
        ", ".join(get_value_list([r.total_pop for r in rew_set_db_rows], no_value_target_pop))
    )
    
    hf_catchment_sheet.cell(
        row_number, 
        1+header_label_to_idx['Catchment Population Total (REW)'],
        ", ".join(get_value_list([r.total_pop for r in rew_set_db_rows], no_value_target_pop))
    )
    
    hf_catchment_sheet.cell(
        row_number, 
        1+header_label_to_idx['Hard to Reach/ Nomadic/ Riverine (REW)'],
        ", ".join(get_value_list([r.problems for r in rew_set_db_rows], no_value_target_pop))
    )
        
    hf_catchment_sheet.cell(
        row_number, 
        1+header_label_to_idx['ANC (REW)'],
        ", ".join(get_value_list([r.anc for r in rew_set_db_rows], no_value_target_pop))
    )
    
    hf_catchment_sheet.cell(
        row_number, 
        1+header_label_to_idx['Family Planning (REW)'],
        ", ".join(get_value_list([r.family_planning for r in rew_set_db_rows], no_value_target_pop))
    )
    
    hf_catchment_sheet.cell(
        row_number, 
        1+header_label_to_idx['Labour & Delivery (REW)'],
        ", ".join(get_value_list([r.labour_delivery for r in rew_set_db_rows], no_value_target_pop))
    )
    
    if rew_set_db_row.is_hf_match:
        
        c = hf_catchment_sheet.cell(
            row_number,
            1+header_label_to_idx['Row Type'],
            "REW Settlement same Health Facility"
        )
        
    else:
        
        c = hf_catchment_sheet.cell(
            row_number,
            1+header_label_to_idx['Row Type'],
            "REW Settlement different Health Facility"
        )
        
        # Color the entire row purple
        purple_fill = PatternFill(start_color="FFA3FF", 
                                  end_color="FFA3FF", fill_type="solid")
        
        last_col_idx = 1+header_label_to_idx['TB/Leprosy services']
        
        # we want inclusive range
        for col_number in range(1, 1+last_col_idx):
            c = hf_catchment_sheet.cell(row_number, col_number)
            
            c.fill = copy(purple_fill)


def fill_in_gmt_no_rew_match(
    hf_catchment_sheet: Worksheet, 
    row_number: int, 
    header_label_to_idx: Dict[str, int]):
    """
    Handle a GMT settlement row that has no REW match
    """
    
    no_value = "Not present in REW"
    
    # get the 1st cell to copy the fill color
    ref_cell = hf_catchment_sheet.cell(row_number, 1)
    
    c = hf_catchment_sheet.cell(
        row_number,
        1+header_label_to_idx['Row Type'],
        "Settlement with no REW Match"
    )
    
    c.fill = copy(ref_cell.fill)
    
    hf_catchment_sheet.cell(
        row_number,
        1+header_label_to_idx['Index'],
        row_number - HF_CATCHMENTS_SHEET_HEADER_ROW - 1
    )
    
    for column_name in [
        'Settlement ID (REW)',
        'Health Facility (REW)',
        'Village/Settlement (REW)',
        'Settlement Type (Urban/ Rural) (REW)',
        'Fixed/Outreach/Mobile (REW)',
        'Type of Immunization Sessions (FS, OS1, OS2, etc.) (REW)',
        'Days of Routine Immunization (fixed post) (REW)',
        'Total Population (REW)',
        'Catchment Population Total (REW)',
        'Hard to Reach/ Nomadic/ Riverine (REW)',
        'ANC (REW)',
        'Family Planning (REW)',
        'Labour & Delivery (REW)',    
    ]:
    
        hf_catchment_sheet.cell(
            row_number, 
            1+header_label_to_idx[column_name],
            no_value)
    
    
    
def modify_gmt_xlsx(conn, lga_name):
    
    conn.set_session(readonly=True)
    cur = conn.cursor()
    tuple_cur = conn.cursor(cursor_factory=psycopg2.extras.NamedTupleCursor) 
    
    workbook, excel_path = setup_clean_workbook(lga_name)
    
    hf_catchment_sheet = workbook.worksheets[4]
    
    assert "HF_catchments" == hf_catchment_sheet.title
        
    assert isinstance(hf_catchment_sheet, Worksheet)
    
    
    # print(header_row)
    
    delete_existing_table(hf_catchment_sheet)
    
    
    assert not isinstance(hf_catchment_sheet, WriteOnlyWorksheet)
    assert isinstance(hf_catchment_sheet, Worksheet)
    assert isinstance(hf_catchment_sheet.max_row, int) 
    
    # 1 based row number, start right before the data, so this should be the 1 based header row
    row_number: int = HF_CATCHMENTS_SHEET_HEADER_ROW
    
    current_fixed_post_row_number = None 
    
    header_label_to_idx: Dict[str,int] = setup_header(hf_catchment_sheet)    
    
    while row_number < hf_catchment_sheet.max_row:
    #for row_number in range(7, hf_catchment_sheet.max_row): # type: ignore
        
        row_number += 1
        
        hf_line = [ cast(Cell,c).value for c in hf_catchment_sheet[row_number]] 
        hf_type = get_text_value(hf_line[header_label_to_idx['Fixed/Outreach/Mobile  (GMT)']])
        
        # If blank we are past the table data
        if hf_type is None:
            break 
        
        # Fix index as we could have inserted rows
        hf_catchment_sheet.cell(
            row_number,
            1+header_label_to_idx['Index'],
            row_number - HF_CATCHMENTS_SHEET_HEADER_ROW - 1
        )
        
        set_name = get_text_value(hf_line[header_label_to_idx['Village/Settlement Name']])
        set_ward = get_expected_text_value(hf_line[header_label_to_idx['Settlement Ward']])
        
        hf_state_lga_ward_name: Tuple[str,str,str,str] = cast(Tuple[str,str,str,str], tuple([
            get_expected_text_value(hf_line[header_label_to_idx[col_label]])
            for col_label in 
            ['HF State', 'HF LGA', 'HF Ward', 'HF Name']
        ]))
        
        # fix mobile colors
        if hf_type == "Mobile":
            fix_mobile_rows(hf_catchment_sheet, set_name, row_number, header_label_to_idx)
                
        if set_name == NA_VALUE or set_name is None:
            
            # Here we update as needed the overall HF row (light blue) or the fixed post one (dark green)
            
            
            if current_fixed_post_row_number is not None:
                # here we want to insert rows
                row_number = insert_unmatched_settlements(tuple_cur, current_fixed_post_row_number, 
                                                          hf_state_lga_ward_name, hf_catchment_sheet, row_number, header_label_to_idx)
                current_fixed_post_row_number = None
                
            if hf_type == 'All':      
                # print("ALL ROW")          
                handle_all_agg_row(cur, hf_state_lga_ward_name, hf_catchment_sheet, row_number, header_label_to_idx)
            elif hf_type == 'Fixed Post':
                handle_fixed_post_agg_row(cur, hf_state_lga_ward_name, hf_catchment_sheet, row_number, header_label_to_idx)
                current_fixed_post_row_number = row_number
            elif hf_type == 'Outreach':
                handle_outreach_agg_row(hf_catchment_sheet, row_number, header_label_to_idx)
            elif hf_type == 'Mobile':
                handle_mobile_agg_row(cur, hf_state_lga_ward_name, hf_catchment_sheet, row_number, header_label_to_idx)
            
            
            
            continue 
        
        if set_name is None or set_name == "":
            break
        
        print(f"Row: {row_number} Looking at [{set_name}] in [{set_ward}]")
        
        assert set_ward is not None 
        sn_matches = get_sn_matches(cur, set_name, set_ward, hf_line, header_label_to_idx)
        
        sn_guid = sn_matches[0][0]
        
        # Do simple cases where we have exactly 1 match
        
        rew_set_db_rows = get_rew_rows(tuple_cur, sn_guid, hf_state_lga_ward_name)
        
        if rew_set_db_rows is None or len(rew_set_db_rows) == 0:
            fill_in_gmt_no_rew_match(hf_catchment_sheet, row_number, header_label_to_idx)
            continue 
        
        # Single line edits
        
        fill_in_rew_inline(hf_catchment_sheet, rew_set_db_rows, row_number, header_label_to_idx)
        
        # if row_number > 30:
        #     break 
        
    # Add the table back
    table = Table(displayName="HF_CatchmentsTable", ref=f'A6:CN{row_number+1}')
    
    # tweak the table styles
    header_fill = PatternFill(start_color="4F81BD", 
                                            end_color="4F81BD", fill_type="solid")
    
    rew_header_fill = PatternFill(
        start_color="B7B7B7", 
        end_color="B7B7B7", 
        fill_type="solid")    
    
    white_font = Font(color="FFFFFF", bold=True)
    
    # set header row styles
    for col_name, col_idx in header_label_to_idx.items():
        c = hf_catchment_sheet.cell(
            row=HF_CATCHMENTS_SHEET_HEADER_ROW, 
            column=1+col_idx)
        if "(REW)" in col_name:
            c.fill = rew_header_fill
        else:
            c.fill = header_fill
        c.font = white_font

    # Add the table to the worksheet
    hf_catchment_sheet.add_table(table)
        
    workbook.save(excel_path)